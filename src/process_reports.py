from datetime import datetime
import decimal
import glob
import operator

from openpyxl import load_workbook

import data.mongo_setup
data.mongo_setup.global_init()

from data.dues import Dues
from data.members import Member
from data.transactions import Transaction, AdminTransaction, Account, AdminAccount


if 0:
    Account.drop_collection()
    settings = {'charity': {'balance':decimal.Decimal(28764.70), 'first_col': 0, 'trans': Transaction, 'acc': Account},
                'admin': {'balance':decimal.Decimal(37369.24), 'first_col': 1, 'trans': AdminTransaction, 'acc': AdminAccount}}
    for account in ('charity', 'admin'):
    # for account in ('charity',):
        acc = settings[account]['acc'](name=account)
        acc.initial_balance = settings[account]['balance']
        files = glob.glob('reports/*.xlsx')
        files.sort()
        for fn in files:
            wb = load_workbook(filename = fn, data_only=True, guess_types=True)
            report_month = fn.split('/')[-1].split('xx')[0]
            sh = wb[account.upper()]
            in_trans = False
            for row in sh.values:
                row = row[settings[account]['first_col']:]
                if all(c is None for c in row):
                    in_trans = False
                if in_trans:
                    if row[0]:
                        date = row[0].date()
                    if row[1] and 'Balance' not in row[1]:
                        for (index, trans_type, op) in ((2, 'payment', operator.sub),
                                                        (3, 'deposit', operator.add)):
                            try:
                                val = row[index].strip()
                            except Exception as e:
                                val = row[index]
                            if val:
                                amt = decimal.Decimal(val)
                                break
                        kwds = {}
                        if account == 'admin':
                            kwds['bar'] = False
                            if any(c in row[1].lower() for c in ('bar ', 'drinks ')):
                                kwds['bar'] = True
                        print(f'{account} {trans_type} on {date.strftime("%y/%m/%d")}: {amt:.2f}. "{row[1]}".')
                        acc.transactions.append(settings[account]['trans'](trans_type=trans_type, trans_date=date, description=row[1], 
                                                                           amount=amt, report_month=report_month, **kwds))
                # find date info
                if row[0] == 'Date':
                    in_trans = True
        acc.save()
        print(f'{acc.current_balance()[1].amount:.2f}')
        if account == 'admin':
            print(acc.bar_values())

if 0:
    wb = load_workbook(filename = 'reports/1809xx.xlsx', data_only=True, guess_types=True)
    sh = wb['CHARITY']
    in_trans = False
    for row in sh.values:
        if all(c is None for c in row):
            in_trans = False
        if in_trans:
            (ln, fn) = row[1].strip().rsplit(' ', 1)
            d = Dues(total=decimal.Decimal(row[0]), discount=decimal.Decimal(row[2] if row[2] else 0), paid=decimal.Decimal(row[3] if row[3] else 0))
            m = Member(first_name=fn, last_name=ln, dues=d)
            m.save()
        if (row[0] == 'Dues') and (row[1] == 'Name'):
            in_trans = True

if 0:
    acc = Account.objects(name='charity').first()
    print(acc.current_balance(month='1807'))
    print(acc.current_balance(month='1808'))
    print(acc.current_balance(month='1809'))
    print(acc.current_balance(month='1810'))
